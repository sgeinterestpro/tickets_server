#!/usr/bin/env python3
# encoding: utf-8

"""
@Time    : 2019/7/9 12:20
@Author  : Sam
@Email   : muumlover@live.com
@Blog    : https://blog.muumlover.com
@Project : tickets_server
@FileName: u_report.py
@Software: PyCharm
@license : (C) Copyright 2019 by muumlover. All rights reserved.
@Desc    :

"""
import logging
from datetime import datetime, timedelta
from io import BytesIO
from typing import List

import openpyxl
import pymongo
from aiohttp.abc import Application
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from base.u_email import EmailSender, EmailContext
from model import Ticket, User, UserInit, TicketCheck, TicketBatch
from unit import date_month_start, date_month_end, date_fmt_conv


def setup_report(app: Application) -> None:
    if 'config' not in app:
        raise Exception('需要初始化配置参数')
    if 'email' not in app:
        raise Exception('需要初始化邮件模块')
    SheetMaker.config = app['config']
    SheetMaker.sport_map = SheetMaker.config.get('ticket', {}).get('sport', {})
    SheetMaker.state_map = SheetMaker.config.get('ticket', {}).get('state', {})
    ReportBase._sender = app['email']

    app['report'] = {
        'ReportUsedDtl': ReportUsedDtl,  # 运动券领用明细日报表
        'ReportUsedDay': ReportUsedDay,  # 运动券领用统计日报表
        'ReportUsedMonth': ReportUsedMonth,  # 运动券领用统计月报表
        'ReportDayCheck': ReportDayCheck,  # 运动券勾稽关系统计表
        'ReportTicketDtl': ReportTicketDtl,  # 运动券发行明细表
        'ReportRaiseDtl': ReportRaiseDtl,  # 运动券发行记录表
    }


def style_title(cell) -> None:
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=16, bold=True)


def style_body(cell):
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))


def style_body_warn(cell):
    cell.fill = PatternFill("solid", fgColor="FFBBBB")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))


def set_merge_content(sheet, row_index, column_index, length, value, pattern=style_body):
    sheet.merge_cells(start_row=row_index, start_column=column_index,
                      end_row=row_index, end_column=column_index + length - 1)
    pattern(sheet.cell(row_index, column_index, value))
    for column_x in range(column_index + 1, column_index + length):
        pattern(sheet.cell(row_index, column_x))


def set_array_content(sheet, row_index, row_data, pattern=style_body):
    for index, data in enumerate(row_data):
        pattern(sheet.cell(row_index, index + 1, data))


def set_field_title(sheet, row_index, field_title):
    for index, (head, width) in enumerate(field_title):
        sheet.column_dimensions[get_column_letter(index + 1)].width = width
        style_body(sheet.cell(row_index, index + 1, head))


def set_large_title(sheet, row_index, column_index, length, value):
    sheet.row_dimensions[row_index].hight = 28.5
    set_merge_content(sheet, row_index, column_index, length, value, pattern=style_title)


def set_export_time(sheet, row_index, column_index, length):
    value = f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}'
    set_merge_content(sheet, row_index, column_index, length, value, pattern=lambda x: x)


def make_body(attach_name):
    return f'''
<style class="fox_global_style"> 
div.fox_html_content {{ line-height: 1.5;}} 
/* 一些默认样式 */ 
blockquote {{ margin-Top: 0px; margin-Bottom: 0px; margin-Left: 0.5em }} 
ol, ul {{ margin-Top: 0px; margin-Bottom: 0px; list-style-position: inside; }} 
p {{ margin-Top: 0px; margin-Bottom: 0px }} 
</style>
<div style="border-bottom:3px solid #d9d9d9;"> 
    <div style="border:1px solid #c8cfda; padding:40px;"> 
        <div> 
            <p>尊敬的用户，您好！<br> <br> </p> 
            <p>您于{datetime.now().strftime("%Y-%m-%d %H:%M")}导出了《{attach_name}》。 <br> 附件为您本次申请导出的报表。</p> 
            <br> <br> 
        </div> 
        <div>本邮件由服务器自动发送</div> 
        <div style="border-top:1px solid #ccc;padding:10px 0;font-size:12px;margin:20px 0;"> 
            票券助手站点：<a href="http://ticket.sge-tech.com">http://ticket.sge-tech.com</a><br> 
        </div> 
    </div> 
</div>
'''


class IncrementCtrl:
    def __init__(self, start=0):
        self._inc_now = start

    @property
    def now(self):
        return self._inc_now

    @property
    def next(self):
        self._inc_now += 1
        return self._inc_now

    def add(self, value=1):
        self._inc_now += value


class ReportBase:
    _title: str = ''
    _sender: EmailSender = None
    _email: EmailContext = None
    _params: dict = {}

    def __init__(self, subject):
        self._title = subject
        self._email = EmailContext(subject)

    async def attach_attach(self):
        return self

    async def send(self, email_addr: List[str]):
        await self.attach_attach()
        if self._email is None:
            return
        await self._sender.send(email_addr, self._email.subject, self._email.message, self._email.attach)
        return self

    @classmethod
    async def to_json(cls, key):
        params = await cls._get_params()
        return {
            'api': key,
            'title': cls._title,
            'params': params,
        }

    @classmethod
    async def _get_params(cls):
        return cls._params


class SheetMaker(object):
    config = None
    sport_map = None
    state_map = None

    @staticmethod
    async def sheet_ticket_dtl(title: str, sheet: Worksheet, batch: str):
        sheet.title = title
        field_title = [
            ('序号', 6),  # 5.38
            ('票券编号', 24),  # 23.38
            ('发布批次', 18),  # 17.38
            ('发布日期', 12),  # 11.38
            ('发布时间', 12),  # 11.38
            ('票券状态', 12),  # 11.38
        ]
        row = IncrementCtrl(0)
        set_large_title(sheet, row.next, 1, len(field_title), title)
        set_export_time(sheet, row.next, 1, len(field_title))
        set_field_title(sheet, row.next, field_title)

        index = IncrementCtrl(0)
        async for ticket in Ticket.find({'batch': batch}):
            set_array_content(sheet, row.next, [
                index.next,  # 序号
                ticket.json_id[:20],  # 票券编号
                batch,  # 发布批次
                ticket['raise_time'].date(),  # 发布日期
                ticket['raise_time'].time(),  # 发布时间
                SheetMaker.state_map.get(ticket.get('state'), '-'),  # 票券状态
            ])

    @staticmethod
    async def sheet_month_count(title: str, sheet: Worksheet, start: str, end: str):
        """
        运动券使用统计月报表
        :param title:
        :param sheet:
        :param start:
        :param end:
        :return:
        """

        async def get_data():
            """
            整理数据
            :return:
            """
            datetime_start = datetime.strptime(start, '%Y-%m')
            datetime_end = datetime.strptime(end, '%Y-%m')
            month_list = []
            month_now = datetime(datetime_start.year, datetime_start.month, 15)
            month_list.append(month_now)
            while month_now < datetime(datetime_end.year, datetime_end.month, 15):
                month_now += timedelta(days=30)
                month_now += timedelta(days=15 - month_now.day)
                month_list.append(month_now)
            _sport_months = []
            _sport_all = {'sports': dict(zip(SheetMaker.sport_map.keys(), [0] * len(SheetMaker.sport_map)))}
            for _month in month_list:
                month_start = date_month_start(_month)
                month_end = date_month_end(_month)
                if datetime.now() < month_start:
                    break
                month_start_str = month_start.strftime('%Y-%m-%d')
                month_end_str = month_end.strftime('%Y-%m-%d')
                month_str = _month.strftime('%Y-%m')
                _sport_month = {'month': month_str, 'start': month_start_str, 'end': month_end_str}
                if month_str not in _sport_months:
                    _sport_month['sports'] = dict(zip(SheetMaker.sport_map.keys(), [0] * len(SheetMaker.sport_map)))
                if datetime.now() <= month_end:
                    _sport_month['un-end'] = True
                async for ticket in Ticket.find({
                    'expiry_date': {'$gte': month_start_str, '$lte': month_end_str}
                }).sort('expiry_date'):
                    if ticket.get('state') == 'verified':
                        _sport_month['sports'][ticket.get('class')] += 1
                        _sport_all['sports'][ticket.get('class')] += 1
                _sport_months.append(_sport_month)
            return _sport_months, _sport_all

        sport_months, sport_all = await get_data()
        sheet.title = title
        field_title = [
            ('年度', 5),
            ('月份', 5),
            *[(x, 7) for x in SheetMaker.sport_map.values()],
            ('合计', 7),
            ('开始日期', 11),
            ('结束日期', 11),
            ('备注', 8),
        ]
        row = IncrementCtrl(0)
        set_large_title(sheet, row.next, 1, len(field_title), title)
        set_export_time(sheet, row.next, 1, len(field_title))
        set_field_title(sheet, row.next, field_title)

        for sport_month in sport_months:
            break_need = bool(sport_month.get('un-end'))  # 当前月计算完就停止
            year, month = sport_month['month'].split('-')
            sport = [
                count for sport, count in sport_month['sports'].items() if sport != 'expired' and sport != 'un-end'
            ]
            note, pattern = ('本月结束后数据可能会发生变动', style_body_warn) if break_need else ('-', style_body)
            set_array_content(sheet, row.next, [
                year,  # 年度
                month,  # 月份
                *sport,  # 运动项目
                sum(sport),  # 合计
                sport_month['start'],  # 开始日期
                sport_month['end'],  # 结束日期
                note,  # 备注
            ], pattern=pattern)
            if break_need:
                break

        # 写入合计
        sport = [count for sport, count in sport_all['sports'].items()]
        set_array_content(sheet, row.next, [
            '合计',  # 合计
            None,  # 合计(被合并的单元格)
            *sport,  # 运动项目
            sum(sport),  # 合计
            '-',  # 开始日期
            '-',  # 结束日期
            '-',  # 备注
        ])
        set_merge_content(sheet, row.now, 1, 2, '合计')

    @staticmethod
    async def sheet_day_count(title: str, sheet: Worksheet, start: str, end: str):
        """
        运动券使用统计日报表
        :param title:
        :param sheet:
        :param start:
        :param end:
        :return:
        """

        async def get_data():
            _sport_days = {}
            _sport_total = dict(zip(SheetMaker.sport_map.keys(), [0] * len(SheetMaker.sport_map)))
            datetime_start = datetime.strptime(start, '%Y-%m-%d')
            datetime_end = datetime.strptime(end, '%Y-%m-%d')
            for i in range((datetime_end - datetime_start).days + 1):
                datetime_now = datetime_start + timedelta(days=i)
                date_now = datetime_now.strftime('%Y-%m-%d')
                _sport_days[date_now] = dict((zip(SheetMaker.sport_map.keys(), [0] * len(SheetMaker.sport_map))))

            cursor = Ticket.find({
                'state': 'verified',
                'expiry_date': {'$gte': start, '$lte': end}
            }).sort('expiry_date')

            async for ticket in cursor:
                date_now = ticket.get('expiry_date', '-')
                _sport_days[date_now][ticket.get('class')] += 1
                _sport_total[ticket.get('class')] += 1
            return _sport_days, _sport_total

        sport_days, sport_total = await get_data()
        sheet.title = title
        field_title = [
            ('日期', 12),
            *[(x, 8) for x in SheetMaker.sport_map.values()],
            ('合计', 8),
            ('备注', 8),
        ]
        row = IncrementCtrl(0)
        set_large_title(sheet, row.next, 1, len(field_title), title)
        set_export_time(sheet, row.next, 1, len(field_title))
        set_field_title(sheet, row.next, field_title)

        for (sport_day, sports) in sport_days.items():
            break_need = bool(datetime.now().strftime('%Y-%m-%d') == sport_day)  # 今日计算完就停止
            sport = [count for sport, count in sports.items() if sport != 'expired']
            note, pattern = ('今日结束后数据可能会发生变动', style_body_warn) if break_need else ('-', style_body)
            set_array_content(sheet, row.next, [
                sport_day,  # 日期
                *sport,  # 运动项目
                sum(sport),  # 合计
                note,  # 备注
            ], pattern=pattern)
            if break_need:
                break

        # 写入合计
        sport = [count for sport, count in sport_total.items()]
        set_array_content(sheet, row.next, [
            '合计',  # 合计
            *sport,  # 运动项目
            sum(sport),  # 合计
            None,  # 备注
        ])

    @staticmethod
    async def sheet_day_dtl(title: str, sheet: Worksheet, date_list: List[str]):
        """
        运动券使用统计日明细报表
        :param title:
        :param sheet:
        :param date_list:
        :return:
        """
        sheet.title = title
        field_title = [
            ('序号', 5.5),  # 4.88
            ('部门', 13),  # 12.38
            ('姓名', 9),  # 8.38
            ('项目', 9),  # 8.38
            ('票券编号', 23),  # 22.38
            ('使用日期', 12),  # 11.38
            ('使用时间', 12),  # 11.38
            ('检票员', 9),  # 8.38
        ]
        row = IncrementCtrl(0)
        set_large_title(sheet, row.next, 1, len(field_title), title)
        # set_large_title(sheet, row.next, 1, len(field_title), date)
        set_export_time(sheet, row.next, 1, len(field_title))
        set_field_title(sheet, row.next, field_title)

        index = IncrementCtrl(0)
        for date in date_list:
            async for ticket in Ticket.find({
                'state': 'verified',
                'expiry_date': {'$gte': date, '$lte': date}
            }).sort('expiry_date'):
                user_init = await UserInit.find_one({'_id': ticket['purchaser']})
                checker_init = await UserInit.find_one({'_id': ticket['checker']})
                set_array_content(sheet, row.next, [
                    index.next,  # 序号
                    (user_init or {}).get('department', '-'),  # 部门
                    (user_init or {}).get('real_name', '-'),  # 姓名
                    SheetMaker.sport_map.get(ticket.get('class'), '-'),  # 项目
                    ticket.json_id[:20],  # 票券编号
                    ticket['check_time'].date(),  # 使用日期
                    ticket['check_time'].time(),  # 使用时间
                    (checker_init or {}).get('real_name', '-'),  # 检票人员
                ])

    @staticmethod
    async def sheet_day_check(title: str, sheet: Worksheet, start: str, end: str):
        sheet.title = title
        field_title = [
            ('统计日期', 21),
            ('票券总量', 10),
            ('未使用量', 10),
            ('已使用量', 10),
        ]
        row = IncrementCtrl(1)
        set_large_title(sheet, row.now, 1, len(field_title), title)
        set_export_time(sheet, row.next, 1, len(field_title))
        set_field_title(sheet, row.next, field_title)

        async for ticket_check in TicketCheck.find({
            'checked_date': {'$gte': start, '$lte': end}
        }).sort('checked_date'):
            set_array_content(sheet, row.next, [
                ticket_check.get('checked_date'),  # 统计日期
                ticket_check.get('all'),  # 票券总量
                ticket_check.get('default'),  # 未使用量
                ticket_check.get('verified'),  # 已使用量
            ])

    @staticmethod
    async def sheet_raise_dtl(title: str, sheet: Worksheet):
        sheet.title = title
        field_title = [
            ('序号', 6),  # 5.38
            ('发布批次', 18),  # 17.38
            ('发布人', 7.5),  # 6.88
            ('发布日期', 12),  # 11.38
            ('发布时间', 12),  # 11.38
            ('数量', 7),  # 6.38
            ('复核人', 7.5),  # 6.88
        ]
        row = IncrementCtrl(1)
        set_large_title(sheet, row.now, 1, len(field_title), title)
        set_export_time(sheet, row.next, 1, len(field_title))
        set_field_title(sheet, row.next, field_title)

        index = IncrementCtrl(0)
        ticket_batch: TicketBatch
        async for ticket_batch in TicketBatch.find({}).sort([('raise_time', pymongo.DESCENDING)]):
            raiser_init = await UserInit.find_one({'_id': ticket_batch['raiser']})
            checker_init = await UserInit.find_one({'_id': ticket_batch['checker']})
            set_array_content(sheet, row.next, [
                index.next,  # 序号
                ticket_batch.json_id,  # 发布批次
                raiser_init['real_name'],  # 发布人
                ticket_batch['raise_time'].date(),  # 发布日期
                ticket_batch['raise_time'].time(),  # 发布时间
                ticket_batch['raise_count'],  # 数量
                checker_init['real_name'] if checker_init else '-',  # 复核人
            ])


class ReportUsedDtl(ReportBase):
    _title: str = '运动券领用明细日报表'
    _params: dict = {
        'date': {
            'title': '导出日期',
            'type': 'day'
        }
    }

    def __init__(self, date: str, *args, **kwargs):
        super().__init__(self._title)  # 初始化 _email
        if not date:
            raise KeyError()
        self.date = date_fmt_conv(date)
        self._attach_name = '{}_{}.xlsx'.format(self._email.subject, date_fmt_conv(self.date, fmt_to="%Y.%m.%d"))
        self._email.message = make_body(self._attach_name)

    async def attach_attach(self):
        attach_data = BytesIO()  # 创建内存IO

        wb = openpyxl.Workbook()  # 创建一个文件对象
        await SheetMaker.sheet_day_dtl('运动券使用统计日明细报表', wb.active, [self.date])  # 写入Sheet
        wb.save(attach_data)  # 写出到IO

        self._email.attach = (self._attach_name, attach_data)
        return self


class ReportUsedDay(ReportBase):
    _title: str = '运动券领用统计日报表'
    _params: dict = {
        'start': {
            'title': '导出起始日期',
            'type': 'day'
        },
        'end': {
            'title': '导出结束日期',
            'type': 'day'
        }
    }

    def __init__(self, start: str, end: str, *args, **kwargs):
        super().__init__(self._title)
        if not start or not end:
            raise KeyError()
        self.start = date_fmt_conv(start)
        self.end = date_fmt_conv(end)
        self._attach_name = '{}_{}-{}.xlsx'.format(self._email.subject,
                                                   date_fmt_conv(self.start, fmt_to="%Y.%m.%d"),
                                                   date_fmt_conv(self.end, fmt_to="%Y.%m.%d"))
        self._email.message = make_body(self._attach_name)

    async def attach_attach(self):
        attach_data = BytesIO()  # 创建内存IO

        wb = openpyxl.Workbook()  # 创建一个文件对象
        await SheetMaker.sheet_day_count(self._title, wb.active, self.start, self.end)  # 输出报表内容
        cursor = Ticket.find({
            'state': 'verified',
            'expiry_date': {'$gte': self.start, '$lte': self.end}
        }).sort('expiry_date')
        sheet = wb.create_sheet('票券明细')
        date_list = []
        async for ticket in cursor:
            if ticket.get('expiry_date', '-') not in date_list:
                date_list.append(ticket.get('expiry_date', '-'))
        await SheetMaker.sheet_day_dtl('票券明细', sheet, date_list)  # 输出报表内容
        wb.save(attach_data)  # 写出到IO

        self._email.attach = (self._attach_name, attach_data)
        return self


class ReportUsedMonth(ReportBase):
    _title: str = '运动券领用统计月报表'
    _params: dict = {
        'start': {
            'title': '导出起始月份',
            'type': 'month'
        },
        'end': {
            'title': '导出结束月份',
            'type': 'month'
        }
    }

    def __init__(self, start: str, end: str, *args, **kwargs):
        super().__init__(self._title)
        if not start or not end:
            raise KeyError()
        self.start = start
        self.end = end
        self._attach_name = '{}_{}-{}.xlsx'.format(self._email.subject,
                                                   date_fmt_conv(self.start, fmt_from='%Y-%m', fmt_to="%Y.%m"),
                                                   date_fmt_conv(self.end, fmt_from='%Y-%m', fmt_to="%Y.%m"))
        self._email.message = make_body(self._attach_name)

    async def attach_attach(self):
        attach_data = BytesIO()  # 创建内存IO

        wb = openpyxl.Workbook()  # 创建一个文件对象
        await SheetMaker.sheet_month_count(self._title, wb.active, self.start, self.end)  # 写入Sheet
        wb.save(attach_data)  # 写出到IO

        self._email.attach = (self._attach_name, attach_data)
        return self


class ReportDayCheck(ReportBase):
    _title: str = '运动券勾稽关系统计表'
    _params: dict = {
        'start': {
            'title': '导出起始日期',
            'type': 'day'
        },
        'end': {
            'title': '导出结束日期',
            'type': 'day'
        }
    }

    def __init__(self, start: str, end: str, *args, **kwargs):
        super().__init__(self._title)
        if not start or not end:
            raise KeyError()
        self.start = date_fmt_conv(start)
        self.end = date_fmt_conv(end)
        self._attach_name = '{}_{}-{}.xlsx'.format(self._email.subject,
                                                   date_fmt_conv(self.start, fmt_to="%Y.%m.%d"),
                                                   date_fmt_conv(self.end, fmt_to="%Y.%m.%d"))
        self._email.message = make_body(self._attach_name)

    async def attach_attach(self):
        attach_data = BytesIO()  # 创建内存IO

        wb = openpyxl.Workbook()  # 创建一个文件对象
        await SheetMaker.sheet_day_check(self._title, wb.active, self.start, self.end)  # 输出报表内容
        wb.save(attach_data)  # 写出到IO

        self._email.attach = (self._attach_name, attach_data)
        return self


class ReportTicketDtl(ReportBase):
    _title: str = '运动券发行明细表'

    @classmethod
    async def _get_params(cls):
        ticket_batch_id_list = []
        ticket_batch: TicketBatch
        async for ticket_batch in TicketBatch.find():
            ticket_batch_id_list.append(ticket_batch.json_id)
        return {
            "batch": {
                "title": "发行批次",
                "values": ticket_batch_id_list
            }
        }

    def __init__(self, batch: str, *args, **kwargs):
        super().__init__(self._title)
        if not batch:
            raise KeyError()
        self.raise_batch = batch
        self._attach_name = '{}_{}.xlsx'.format(self._email.subject, batch)
        self._email.message = make_body(self._attach_name)

    async def attach_attach(self):
        attach_data = BytesIO()  # 创建内存IO

        wb = openpyxl.Workbook()  # 创建一个文件对象
        await SheetMaker.sheet_ticket_dtl(self._title, wb.active, self.raise_batch)  # 输出报表内容
        wb.save(attach_data)  # 写出到IO

        self._email.attach = (self._attach_name, attach_data)
        return self


class ReportRaiseDtl(ReportBase):
    _title: str = '运动券发行记录表'

    def __init__(self, *args, **kwargs):
        super().__init__(self._title)
        self._attach_name = '{}.xlsx'.format(self._email.subject)
        self._email.message = make_body(self._attach_name)

    async def attach_attach(self):
        attach_data = BytesIO()  # 创建内存IO

        wb = openpyxl.Workbook()  # 创建一个文件对象
        await SheetMaker.sheet_raise_dtl(self._title, wb.active)  # 输出报表内容
        wb.save(attach_data)  # 写出到IO

        self._email.attach = (self._attach_name, attach_data)
        return self


if __name__ == '__main__':
    from sys import stdout

    logging.basicConfig(
        format='%(levelname)s: %(asctime)s [%(filename)s:%(lineno)d] %(message)s',
        level=logging.NOTSET,
        stream=stdout)

    import pathlib
    import asyncio
    from motor.motor_asyncio import AsyncIOMotorClient
    from base.u_email import EmailSender, EmailSender, EmailSender, EmailSender, EmailSender
    from config import load_config
    from urllib.parse import quote_plus
    from model import Model

    uri = 'mongodb://'
    uri += '{}:{}'.format(quote_plus('127.0.0.1'), quote_plus('27017'))
    client = AsyncIOMotorClient(uri)
    Model._db = client.get_database('ticket_test')
    ReportBase._sender = EmailSender
    SheetMaker.config = load_config(str(pathlib.Path('../..') / 'config' / 'polls.yaml'))
    SheetMaker.sport_map = SheetMaker.config.get('ticket', {}).get('sport', {})
    SheetMaker.state_map = SheetMaker.config.get('ticket', {}).get('state', {})


    async def test():
        with open('ReportUsedDtl.xlsx', 'wb') as xls:
            attach = (await ReportUsedDtl('2019-11-05').attach_attach())._email.attach[1]
            xls.write(attach.getvalue())
        with open('ReportUsedDay.xlsx', 'wb') as xls:
            attach = (await ReportUsedDay('2019-01-01', '2019-12-31').attach_attach())._email.attach[1]
            xls.write(attach.getvalue())
        with open('ReportUsedMonth.xlsx', 'wb') as xls:
            attach = (await ReportUsedMonth('2019-01-01', '2019-12-31').attach_attach())._email.attach[1]
            xls.write(attach.getvalue())
        with open('ReportDayCheck.xlsx', 'wb') as xls:
            attach = (await ReportDayCheck('2019-01-01', '2019-12-31').attach_attach())._email.attach[1]
            xls.write(attach.getvalue())
        with open('ReportTicketDtl.xlsx', 'wb') as xls:
            attach = (await ReportTicketDtl('201911081BE64582').attach_attach())._email.attach[1]
            xls.write(attach.getvalue())
        with open('ReportRaiseDtl.xlsx', 'wb') as xls:
            attach = (await ReportRaiseDtl().attach_attach())._email.attach[1]
            xls.write(attach.getvalue())
        # with open('ReportUsedSportMonth.xlsx', 'wb') as xls:
        #     attach = await ReportUsedSportMonth('2019-01-01', '2019-12-31').get_attachs()
        #     xls.write(attach[1].getvalue())


    loop = asyncio.get_event_loop()
    loop.run_until_complete(test())
