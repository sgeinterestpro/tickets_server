#!/usr/bin/env python3
# encoding: utf-8

"""
@Time    : 2019/7/9 12:20
@Author  : Sam
@Email   : muumlover@live.com
@Blog    : https://blog.muumlover.com
@Project : tickets_server
@FileName: u_report.py
@Software: PyCharm
@license : (C) Copyright 2019 by muumlover. All rights reserved.
@Desc    :

"""
import logging
from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from u_email import EmailSender
from unit import date_month_start, date_month_end


def setup_report(app):
    ReportBase.sender = app['email']
    ReportBase.db = app['db']
    ReportBase.sport_map = app['config'].get('ticket', {}).get('sport', {})
    ReportBase.state_map = app['config'].get('ticket', {}).get('state', {})
    app['report'] = {
        'ReportBase': ReportBase,
        'ReportCheckLogFlow': ReportCheckLogFlow,
        'ReportUsedDtl': ReportUsedDtl,
        'ReportUsedDay': ReportUsedDay,
        'ReportUsedMonth': ReportUsedMonth,
        'ReportUsedSportMonth': ReportUsedSportMonth
    }


def style_title(cell):
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=16, bold=True)


def style_body(cell):
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))


def style_body_warn(cell):
    cell.fill = PatternFill("solid", fgColor="FFBBBB")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))


def date_show(date, fmt):
    return datetime.strptime(date, "%Y-%m-%d").strftime(fmt)


class ReportBase:
    sender: EmailSender = None
    db = None
    config = None
    sport_map = None
    state_map = None

    _email_subject = '报表导出'
    _email_attach = _email_subject

    @property
    def _mail_msg(self):
        return f'''您好，附件为您本次申请导出的{self._email_attach}。'''

    async def get_attachs(self):
        wb = openpyxl.Workbook()  # 创建一个文件对象
        await self.write_wb(wb)
        output = BytesIO()  # 创建内存IO
        wb.save(output)  # 写出到IO
        return f'{self._email_attach}.xlsx', output

    async def write_wb(self, wb):
        pass

    async def send(self, email_addr):
        attachs = await self.get_attachs()
        await self.sender.send(email_addr, self._email_subject, self._mail_msg, attachs)

    async def sheet_day_count(self, sheet, date_start, date_end):
        sheet.row_dimensions[1].hight = 28.5
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.sport_map) + 3)
        style_title(sheet.cell(1, 1, '票券使用统计日报表'))

        # 写入数据列标题
        now_column = 1
        sheet.column_dimensions[get_column_letter(now_column)].width = 12
        style_body(sheet.cell(2, now_column, '日期'))
        now_column += 1
        for sport in self.sport_map.values():
            sheet.column_dimensions[get_column_letter(now_column)].width = 8.38
            style_body(sheet.cell(2, now_column, sport))
            now_column += 1
        sheet.column_dimensions[get_column_letter(now_column)].width = 8.38
        style_body(sheet.cell(2, now_column, '合计'))
        now_column += 1
        sheet.column_dimensions[get_column_letter(now_column)].width = 8.38
        style_body(sheet.cell(2, now_column, '备注'))

        # 填充数据
        sport_days = {}
        datetime_start = datetime.strptime(date_start, '%Y-%m-%d')
        datetime_end = datetime.strptime(date_end, '%Y-%m-%d')
        for i in range((datetime_end - datetime_start).days + 1):
            datetime_now = datetime_start + timedelta(days=i)
            date_now = datetime_now.strftime('%Y-%m-%d')
            sport_days[date_now] = dict((zip(self.sport_map.keys(), [0] * len(self.sport_map))))

        cursor = self.db.ticket.find({
            'state': 'verified',
            'expiry_date': {'$gte': date_start, '$lte': date_end}
        }).sort('expiry_date')

        async for ticket_doc in cursor:
            date_now = ticket_doc.get('expiry_date', '-')
            sport_days[date_now][ticket_doc.get('class')] += 1

        for sport_day_i, (sport_day, sports) in enumerate(sport_days.items()):
            now_row = sport_day_i + 3
            now_column = 1
            style_body(sheet.cell(now_row, now_column, sport_day))  # 日期
            now_column += 1
            count_all = 0
            for sport, count in sports.items():
                if sport == 'expired':
                    continue
                style_body(sheet.cell(now_row, now_column, count))  # 运动项目
                count_all += count
                now_column += 1
            style_body(sheet.cell(now_row, now_column, count_all))  # 合计
            now_column += 1
            note = None
            if datetime.now().strftime('%Y-%m-%d') == sport_day:
                note = '今日结束后数据可能会发生变动'
            style_body(sheet.cell(now_row, now_column, note))  # 备注

    async def sheet_day_dtl(self, sheet, date):
        sheet.row_dimensions[1].hight = 28.5
        sheet.merge_cells('A1:G1')
        style_title(sheet.cell(1, 1, '票券使用明细表'))
        # 写入报表日期
        sheet.row_dimensions[2].hight = 28.5
        sheet.merge_cells('A2:G2')
        style_title(sheet.cell(2, 1, date))
        # 写入数据列标题
        for index, (head, width) in enumerate([
            ('序号', 5),
            ('部门', 10),
            ('姓名', 6.5),
            ('项目', 6.5),
            ('票券编号', 23),
            ('领取时间', 21),
            ('使用时间', 21)
        ]):
            sheet.column_dimensions[get_column_letter(index + 1)].width = width
            style_body(sheet.cell(3, index + 1, head))
        # 填充数据
        cursor = self.db.ticket.find({
            'state': 'verified',
            'expiry_date': {'$gte': date, '$lte': date}
        }).sort('expiry_date')
        index = 0
        offset = 0
        async for ticket_doc in cursor:
            now_row = index + offset + 4
            user = await self.db.user.find_one({'_id': ticket_doc['purchaser']})
            user_init = await self.db.user_init.find_one({'_id': (user or {}).get('init_id')})
            style_body(sheet.cell(now_row, 1, index + 1))  # 序号
            style_body(sheet.cell(now_row, 2, (user_init or {}).get('department', '-')))  # 部门
            style_body(sheet.cell(now_row, 3, (user_init or {}).get('real_name', '-')))  # 姓名
            style_body(sheet.cell(now_row, 4, self.sport_map.get(ticket_doc.get('class'), '-')))  # 项目
            style_body(sheet.cell(now_row, 5, str(ticket_doc.get('_id', '-'))[:20]))  # 票券编号
            style_body(sheet.cell(now_row, 6, ticket_doc.get('purch_time', '-')))  # 领取时间
            style_body(sheet.cell(now_row, 7, ticket_doc.get('check_time', '-')))  # 使用时间
            index += 1

    async def sheet_ticket_dtl(self, sheet, date):
        sheet.row_dimensions[1].hight = 28.5
        sheet.merge_cells('A1:G1')
        style_title(sheet.cell(1, 1, '票券使用明细表'))
        # 写入报表日期
        sheet.row_dimensions[2].hight = 28.5
        sheet.merge_cells('A2:G2')
        style_title(sheet.cell(2, 1, date))
        # 写入数据列标题
        for index, (head, width) in enumerate([
            ('序号', 5),
            ('部门', 10),
            ('姓名', 6.5),
            ('项目', 6.5),
            ('票券编号', 23),
            ('领取时间', 21),
            ('票券状态', 8.5)
        ]):
            sheet.column_dimensions[get_column_letter(index + 1)].width = width
            style_body(sheet.cell(3, index + 1, head))
        # 填充数据
        cursor = self.db.ticket.find({
            'expiry_date': {'$gte': date, '$lte': date}
        }).sort('expiry_date')
        index = 0
        offset = 0
        async for ticket_doc in cursor:
            now_row = index + offset + 4
            user = await self.db.user.find_one({'_id': ticket_doc['purchaser']})
            user_init = await self.db.user_init.find_one({'_id': (user or {}).get('init_id')})
            style_body(sheet.cell(now_row, 1, index + 1))  # 序号
            style_body(sheet.cell(now_row, 2, (user_init or {}).get('department', '-')))  # 部门
            style_body(sheet.cell(now_row, 3, (user_init or {}).get('real_name', '-')))  # 姓名
            style_body(sheet.cell(now_row, 4, self.sport_map.get(ticket_doc.get('class'), '-')))  # 项目
            style_body(sheet.cell(now_row, 5, str(ticket_doc.get('_id', '-'))[:20]))  # 票券编号
            style_body(sheet.cell(now_row, 6, ticket_doc.get('purch_time', '-')))  # 领取时间
            style_body(sheet.cell(now_row, 7, self.state_map.get(ticket_doc.get('state'), '-')))  # 票券状态
            index += 1

    async def sheet_day_check(self, sheet, date):
        sheet.row_dimensions[1].hight = 28.5
        sheet.merge_cells('A1:G1')
        style_title(sheet.cell(1, 1, '票券使用明细表'))
        # 写入报表日期
        sheet.row_dimensions[2].hight = 28.5
        sheet.merge_cells('A2:G2')
        style_title(sheet.cell(2, 1, date))
        # 写入数据列标题
        for index, (head, width) in enumerate([
            ('序号', 5),
            ('部门', 10),
            ('姓名', 6.5),
            ('项目', 6.5),
            ('票券编号', 23),
            ('领取时间', 21),
            ('票券状态', 8.5)
        ]):
            sheet.column_dimensions[get_column_letter(index + 1)].width = width
            style_body(sheet.cell(3, index + 1, head))
        # 填充数据
        cursor = self.db.ticket.find({
            'expiry_date': {'$gte': date, '$lte': date}
        }).sort('expiry_date')
        index = 0
        offset = 0
        async for ticket_doc in cursor:
            now_row = index + offset + 4
            user = await self.db.user.find_one({'_id': ticket_doc['purchaser']})
            user_init = await self.db.user_init.find_one({'_id': (user or {}).get('init_id')})
            style_body(sheet.cell(now_row, 1, index + 1))  # 序号
            style_body(sheet.cell(now_row, 2, (user_init or {}).get('department', '-')))  # 部门
            style_body(sheet.cell(now_row, 3, (user_init or {}).get('real_name', '-')))  # 姓名
            style_body(sheet.cell(now_row, 4, self.sport_map.get(ticket_doc.get('class'), '-')))  # 项目
            style_body(sheet.cell(now_row, 5, str(ticket_doc.get('_id', '-'))[:20]))  # 票券编号
            style_body(sheet.cell(now_row, 6, ticket_doc.get('purch_time', '-')))  # 领取时间
            style_body(sheet.cell(now_row, 7, self.state_map.get(ticket_doc.get('state'), '-')))  # 票券状态
            index += 1


class ReportCheckLogFlow(ReportBase):
    _email_subject = '活动券使用记录流水表'

    async def write_wb(self, wb):
        # 创建一个sheet对象
        sheet = wb.active
        # 写入文件标题
        for index, (head, width) in enumerate([
            ('序号', 4.63),
            ('项目', 8),
            ('票券编号', 23.25),
            ('领取时间', 21),
            ('领取人员', 10),
            ('使用时间', 21),
            ('检票人员', 10)
        ]):
            sheet.column_dimensions[get_column_letter(index + 1)].width = width
            style_body(sheet.cell(3, index + 1, head))

        style_title(sheet.cell(1, 4, self._email_subject))
        cursor = self.db.ticket_log.find({})
        index = 0
        async for ticket_log in cursor:
            if 'checked' != ticket_log['option']:
                continue

            ticket_doc = await self.db.ticket.find_one({'_id': ticket_log['ticket_id'], })
            purchaser = await self.db.user.find_one({'_id': ticket_doc['purchaser'], })
            checker = await self.db.user.find_one({'_id': ticket_doc['checker'], })

            style_body(sheet.cell(index + 4, 1, index + 1))  # 序号
            style_body(sheet.cell(index + 4, 2, ticket_doc['class']))  # 项目
            style_body(sheet.cell(index + 4, 3, ticket_doc['_id']))  # 票券编号
            style_body(sheet.cell(index + 4, 4, ticket_doc['purch_time']))  # 领取时间
            style_body(sheet.cell(index + 4, 5, purchaser['realName']))  # 领取人员
            style_body(sheet.cell(index + 4, 6, ticket_doc['check_time']))  # 使用时间
            style_body(sheet.cell(index + 4, 7, checker['realName']))  # 检票人员

            index += 1


class ReportUsedDtl(ReportBase):
    _email_subject = '领用登记明细表'
    date_start = None
    date_end = None

    def __init__(self, date):
        super().__init__()
        self.date = date
        self._email_attach = '{}_{}'.format(
            self._email_subject,
            date_show(self.date, "%Y.%m.%d"),
        )

    async def write_wb(self, wb):
        sheet = wb.active  # 获取sheet对象
        await self.sheet_day_dtl(sheet, self.date)  # 写入Sheet


class ReportUsedDay(ReportBase):
    _email_subject = '票券使用统计日报表'
    date_start = None
    date_end = None

    def __init__(self, date_start, date_end):
        super().__init__()
        self.date_start = date_start
        self.date_end = date_end
        self._email_attach = '{}_{}-{}'.format(
            self._email_subject,
            date_show(self.date_start, "%Y.%m.%d"),
            date_show(self.date_end, "%Y.%m.%d")
        )

    async def write_wb(self, wb):
        sheet = wb.active  # 使用默认的Sheet
        sheet.title = '统计'  # 写入Sheet标题
        await self.sheet_day_count(sheet, self.date_start, self.date_end)  # 输出报表内容
        cursor = self.db.ticket.find({
            'state': 'verified',
            'expiry_date': {'$gte': self.date_start, '$lte': self.date_end}
        }).sort('expiry_date')
        date_now = self.date_end
        async for ticket_doc in cursor:
            if date_now != ticket_doc.get('expiry_date', '-'):
                date_now = ticket_doc.get('expiry_date', '-')
                sheet = wb.create_sheet(date_now)  # 创建一个sheet对象
                await self.sheet_day_dtl(sheet, date_now)  # 输出报表内容


class ReportUsedMonth(ReportBase):
    _email_subject = '票券使用统计月度报表'
    date_start = None
    date_end = None

    def __init__(self, date_start, date_end):
        super().__init__()
        self.date_start = date_start
        self.date_end = date_end
        self._email_attach = '{}_{}-{}'.format(
            self._email_subject,
            date_show(self.date_start, "%Y.%m"),
            date_show(self.date_end, "%Y.%m")
        )

    async def write_wb(self, wb):
        other = 5

        # 填充数据
        date_start = datetime.strptime(self.date_start, '%Y-%m-%d')
        date_end = datetime.strptime(self.date_end, '%Y-%m-%d')
        month_list = []
        month_now = datetime(date_start.year, date_start.month, 15)
        month_list.append(month_now)
        while month_now < datetime(date_end.year, date_end.month, 15):
            month_now += timedelta(days=30)
            month_now += timedelta(days=15 - month_now.day)
            month_list.append(month_now)
        sport_months = []
        sport_all = {
            'sports': dict(zip(self.sport_map.keys(), [0] * len(self.sport_map)))
        }
        for month in month_list:
            month_start = date_month_start(month)
            month_end = date_month_end(month)
            if datetime.now() < month_start:
                break

            month_start_str = month_start.strftime('%Y-%m-%d')
            month_end_str = month_end.strftime('%Y-%m-%d')
            month_str = month.strftime('%Y-%m')
            sport_month = {'month': month_str, 'start': month_start_str, 'end': month_end_str}
            if month_str not in sport_months:
                sport_month['sports'] = dict(zip(self.sport_map.keys(), [0] * len(self.sport_map)))
            if datetime.now() <= month_end:
                sport_month['un-end'] = True
            cursor = self.db.ticket.find({
                'expiry_date': {'$gte': month_start_str, '$lte': month_end_str}
            }).sort('expiry_date')
            async for ticket_doc in cursor:
                if ticket_doc.get('state') == 'verified':
                    sport_month['sports'][ticket_doc.get('class')] += 1
                    sport_all['sports'][ticket_doc.get('class')] += 1
            sport_months.append(sport_month)

        # 创建一个sheet对象
        sheet = wb.active
        # 写入报表标题
        now_row = 1
        sheet.row_dimensions[now_row].hight = 28.5
        sheet.merge_cells(start_row=now_row, start_column=1, end_row=now_row, end_column=len(self.sport_map) + other)
        style_title(sheet.cell(now_row, 1, self._email_subject))
        now_row += 1
        sheet.row_dimensions[now_row].hight = 28.5
        sheet.merge_cells(start_row=now_row, start_column=1, end_row=now_row, end_column=len(self.sport_map) + other)
        sheet.cell(now_row, 1, f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
        now_row += 1
        # 写入数据列标题
        now_column = 1
        for text, width in [('年度', 5), ('月份', 5), ('开始日期', 11), ('结束日期', 11)]:
            sheet.column_dimensions[get_column_letter(now_column)].width = width
            style_body(sheet.cell(now_row, now_column, text))
            now_column += 1
        for sport in self.sport_map.values():
            sheet.column_dimensions[get_column_letter(now_column)].width = 7
            style_body(sheet.cell(now_row, now_column, sport))
            now_column += 1
        for text, width in [('合计', 7)]:
            sheet.column_dimensions[get_column_letter(now_column)].width = width
            style_body(sheet.cell(now_row, now_column, text))
            now_column += 1
        now_row += 1
        # 写入内容数据
        for sport_month_i, sport_month in enumerate(sport_months):
            body = style_body_warn if sport_month.get('un-end') else style_body
            year, month = sport_month['month'].split('-')
            now_column = 1
            body(sheet.cell(now_row, now_column, year))  # 年度
            now_column += 1
            body(sheet.cell(now_row, now_column, month))  # 月份
            now_column += 1
            body(sheet.cell(now_row, now_column, sport_month['start']))  # 开始日期
            now_column += 1
            body(sheet.cell(now_row, now_column, sport_month['end']))  # 结束日期
            now_column += 1
            count_all = 0
            for sport, count in sport_month['sports'].items():
                if sport == 'expired' or sport == 'un-end':
                    continue
                body(sheet.cell(now_row, now_column, count))  # 运动项目
                count_all += count
                now_column += 1
            body(sheet.cell(now_row, now_column, count_all))  # 合计
            now_row += 1
        # 写入合计
        sheet.merge_cells(start_row=now_row, start_column=1, end_row=now_row, end_column=4)
        now_column = 1
        style_body(sheet.cell(now_row, now_column, '合计'))  # 合计
        now_column += 1
        style_body(sheet.cell(now_row, now_column))  # 合计
        now_column += 1
        style_body(sheet.cell(now_row, now_column))  # 合计
        now_column += 1
        style_body(sheet.cell(now_row, now_column))  # 合计
        now_column += 1
        count_all = 0
        for sport, count in sport_all['sports'].items():
            style_body(sheet.cell(now_row, now_column, count))  # 运动项目
            count_all += count
            now_column += 1
        style_body(sheet.cell(now_row, now_column, count_all))  # 合计


class ReportUsedSportMonth(ReportBase):
    _email_subject = '活动券分项领用月报表'
    date = None

    def __init__(self, date):
        super().__init__()
        self.date = date
        self._email_attach = '{}_{}'.format(
            self._email_subject,
            date_show(self.date, "%Y.%m")
        )

    async def write_wb(self, wb):
        not_end = False
        month_start = date_month_start(datetime.strptime(self.date, '%Y-%m-%d'))
        month_end = date_month_end(datetime.strptime(self.date, '%Y-%m-%d'))
        if month_start < datetime.now() < month_end:
            not_end = True
        month_start = month_start.strftime('%Y-%m-%d')
        month_end = month_end.strftime('%Y-%m-%d')

        # 创建一个sheet对象
        sheet = wb.active
        # 写入报表标题
        merge_column = 5
        if not_end:
            merge_column = 6
        sheet.row_dimensions[1].hight = 28.5
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_column)
        style_title(sheet.cell(1, 1, self._email_subject))
        # 写入报表日期
        sheet.row_dimensions[2].hight = 28.5
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_column)
        date_text = f'{month_start}至{month_end}'
        if not_end:
            date_text += f'【期中{datetime.now().strftime("%Y-%m-%d")}导出数据】'
        style_title(sheet.cell(2, 1, date_text))
        # 写入数据列标题
        for index, (head, width) in enumerate([
            ('序号', 8),
            ('项目', 12),
            ('本期领取', 18),
            ('本期过期作废', 18),
            ('本期实际使用', 18)
        ]):
            sheet.column_dimensions[get_column_letter(index + 1)].width = width
            style_body(sheet.cell(3, index + 1, head))
        if not_end:
            sheet.column_dimensions[get_column_letter(6)].width = 18
            style_body(sheet.cell(3, 6, '本期暂未使用'))

        # 填充数据
        sports = {}  # dict(zip(self.sport_map.keys(), [dict([('expired', 1), ('verified', 2)])] * len(self.sport_map)))
        for sport_key in self.sport_map.keys():
            sports[sport_key] = {'expired': 0, 'verified': 0, 'valid': 0}
        cursor = self.db.ticket.find({
            'expiry_date': {'$gte': month_start, '$lte': month_end}
        }).sort('expiry_date')
        async for ticket_doc in cursor:
            if ticket_doc.get('state') == 'verified':
                sports[ticket_doc.get('class')]['verified'] += 1
            elif ticket_doc.get('state') == 'expired':
                sports[ticket_doc.get('class')]['expired'] += 1
            else:
                sports[ticket_doc.get('class')]['valid'] += 1
        index = 0
        for sport, value in sports.items():
            now_row = index + 4
            style_body(sheet.cell(now_row, 1, index + 1))  # 序号
            style_body(sheet.cell(now_row, 2, self.sport_map.get(sport, '-')))  # 项目
            style_body(sheet.cell(now_row, 3, value['expired'] + value['verified']))  # 本期领取
            style_body(sheet.cell(now_row, 4, value['expired']))  # 本期过期作废
            style_body(sheet.cell(now_row, 5, value['verified']))  # 本期实际使用
            if not_end:
                style_body(sheet.cell(now_row, 6, value['valid']))  # 本期暂未使用
            index += 1


if __name__ == '__main__':
    from sys import stdout

    logging.basicConfig(
        format='%(levelname)s: %(asctime)s [%(filename)s:%(lineno)d] %(message)s',
        level=logging.NOTSET,
        stream=stdout)

    import pathlib
    import asyncio
    from motor.motor_asyncio import AsyncIOMotorClient
    from u_email import EmailSender, EmailSender, EmailSender, EmailSender, EmailSender
    from config import load_config
    from urllib.parse import quote_plus

    uri = 'mongodb://'
    uri += '{}:{}'.format(quote_plus('127.0.0.1'), quote_plus('27017'))
    client = AsyncIOMotorClient(uri)

    ReportBase.db = client.get_database('ticket')
    ReportBase.sender = EmailSender
    ReportBase.config = load_config(str(pathlib.Path('..') / 'config' / 'polls.yaml'))


    async def test():
        with open('ReportUsedDtl.xlsx', 'wb') as xls:
            attachs = await ReportUsedDtl('2019-07-19').get_attachs()
            xls.write(attachs[1].getvalue())
        with open('ReportUsedDay.xlsx', 'wb') as xls:
            attachs = await ReportUsedDay('2019-07-11', '2019-08-19').get_attachs()
            xls.write(attachs[1].getvalue())
        with open('ReportUsedMonth.xlsx', 'wb') as xls:
            attachs = await ReportUsedMonth('2019-03-11', '2019-08-19').get_attachs()
            xls.write(attachs[1].getvalue())
        with open('ReportUsedSportMonth.xlsx', 'wb') as xls:
            attachs = await ReportUsedSportMonth('2019-07-19').get_attachs()
            xls.write(attachs[1].getvalue())


    loop = asyncio.get_event_loop()
    loop.run_until_complete(test())
